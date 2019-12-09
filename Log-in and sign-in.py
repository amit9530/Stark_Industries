"""
Created on Sun Dec  8 10:54:33 2019

@author: Amit
"""
import pandas as pd
from openpyxl import load_workbook


write = load_workbook(filename = "D:\\Project\\UsersList.xlsx")
sheet = write.active
UsersList = pd.read_excel("D:\\Project\\UsersList.xlsx", "Sheet1")


print("Welcome...")
welcome = input("Do you have an acount? y/n: ")


if welcome == "n" or welcome == "N":
    count=0
    for i in UsersList['id']:
        count+=1
    count+=1
    while True:
        username = int(input("Enter a username: "))
        flag = False
        for Id in UsersList['id']:
            if username == Id:
                flag = True
        if flag == False:
            IdCell = sheet.cell(row=count+1, column=1)
            IdCell.value = username
            password = int(input("Enter a password: "))
            PasswordCell=sheet.cell(row=count+1, column=2)
            PasswordCell.value=password
            print("Type 1 for player")
            print("Type 2 for parent")
            print("Type 3 for proffesional")
            usertype = int(input("Enter the user type: "))
            while usertype!=1 and usertype!=2 and usertype!=3:
                usertype = int(input("Wrong input, try again: "))
            TypeCell=sheet.cell(row=count+1, column=3)
            TypeCell.value=usertype
            break
        print("ID already exist")
    write.save(filename = "D:\\Project\\UsersList.xlsx")
    if usertype==1:
        print("Welcome to the Player Menu")
        PlayerMenu()
    elif usertype==2:
        print("Welcome to the Parent Menu")
        ParentMenu()
    elif usertype==3:
        print("Welcome to the Professional Menu")
        ProfessionalMenu()


elif welcome == "y" or welcome == "Y":
    while True:
        username = int(input("Enter a username: "))
        i =0;
        for row in sheet.rows:
            i = i+ 1;
            for cell in row:
                if(cell.value == username):         
                    line=i
        flag = True
        for Id in UsersList['id']:
            if username == Id:
                flag = False
        if flag == False:
            print(line)
            while True:
                password=int(input("Enter a password: "))
                flag=False
                if password!=UsersList['password'][line-2]:
                    flag=True
                if flag==False:
                    if UsersList['type'][line-2]==1:
                        print("Welcome to the Player Menu")
                        PlayerMenu()
                    elif UsersList['type'][line-2]==2:
                        print("Welcome to the Parent Menu")
                        ParentMenu()
                    elif UsersList['type'][line-2]==3:
                        print("Welcome to the Professional Menu")
                        ProfessionalMenu()
                print("Wrong password, try again")
            break
        print("ID not exist in the system")
