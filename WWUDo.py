import pandas as pd
import xlsxwriter
import numpy as np
import functools
import xlrd
from openpyxl import load_workbook
import sys
import time
from random import seed
from random import randint
from random import sample


# -----------------------------------

# Elior Function's
# --------------------------------------------------
def View_Skip(kid_id):  # print the question from last game if skip from "User_db"
    Player_db = pd.read_excel('Player_db.xlsx')
    kid = Player_db.loc[Player_db.ID == kid_id]
    for index, rows in kid.iterrows():
        my_list = [rows.Q1, rows.A1, rows.Q2, rows.A2, rows.Q3, rows.A3, rows.Q4, rows.A4, rows.Q5, rows.A5]
    if "s" not in my_list:
        print("No questions were skipped")
    else:
        i = 0
        while i < len(my_list):
            if my_list[i] == "s":
                print(my_list[i - 1])
            i = i + 1
    # Unit_Test
    if kid_id == (int((kid['ID']))):
        return 0
    else:
        return 1


# --------------------------------------------------

def Add_Kid(parent_id):  # Get kid and parent id and write the parent id in "Parent" in "Player_db"
    Player_db = pd.read_excel('Player_db.xlsx')
    kid_id = int(input("Please enter kid id"))
    Player_db.loc[Player_db.ID == kid_id, 'Parent'] = parent_id
    writer = pd.ExcelWriter('Player_db.xlsx', engine='xlsxwriter')
    Player_db.to_excel(writer)
    writer.save()
    # Unit_Test
    parent = Player_db.loc[Player_db.Parent == parent_id]
    if parent_id == int(parent['Parent']):
        return 0
    else:
        return 1


# --------------------------------------------------

def View_Kid(parent_id):  # Get parent id and print all the kids that belong to the parent id from "Player_db"
    Player_db = pd.read_excel('Player_db.xlsx')
    kids = Player_db.loc[Player_db.Parent == parent_id]
    print(int(kids['ID']))
    # Unit_Test
    parent = Player_db.loc[Player_db.Parent == parent_id]
    if parent_id == int(parent['Parent']):
        return 0
    else:
        return 1


# --------------------------------------------------

def Print_Login_Count():  # Get kid id and print login count from "Player_db"
    Player_db = pd.read_excel('Player_db.xlsx')
    kid_id = int(input("Please enter kid id: "))
    kid = Player_db.loc[Player_db.ID == kid_id]
    print(int(kid['Login_count']))
    if kid_id == (int((kid['ID']))):
        return 0
    else:
        return 1


# --------------------------------------------------

def Example_Game():  # play game for example to Understand how to play the game
    x = 1
    print("Example Game\n Choose answer 1|2|3:")
    while x < 6:
        print("Question:", x)
        print("Answer 1")
        print("Answer 2")
        print("Answer 3")
        user_input = input("Enter answer:")
        if user_input == '1' or user_input == '2' or user_input == '3':
            print("You choose answer : ", user_input)
        else:
            print("*** Wrong  answer! ***\n*** Choose Only  1 | 2 | 3 ***")
            x = x - 1
        x = x + 1
    return 1


# -----------------------------------

def Print_Grades(id):
    """ function gets id of player and prints all grades from the player's data base"""
    player_db = '{0}.xlsx'.format(id)
    grades = pd.read_excel(player_db)
    print('Printing grades: ')
    for index, row in grades.iterrows():
        print('Grade: {0}, date: {1}'.format(row['Grade'], row['Date']))


def Most_Mistakes():
    """ function searches in data base for the question with most mistakes and prints it"""
    questions_db = pd.read_excel('Question_db_new.xlsx')
    max_mistakes = 0
    for mistake in questions_db['Mistakes']:
        if mistake > max_mistakes:
            max_mistakes = mistake
    max_index = (questions_db.index[questions_db['Mistakes'] == max_mistakes].tolist())[0]
    print('The question with the most mistakes is {0}'.format(questions_db.loc[max_index]['Question']))
    print('This question has been answered wrong {0} times.'.format(max_mistakes))


def Delete_User():
    """ function deletes user from player db and users db"""
    id = int(input('Please enter the ID of the user to delete: '))
    # search and delete from players data base
    players = pd.read_excel('Player_db.xlsx')
    for player_id in players['ID']:
        if id == player_id:
            player_index = (players.index[players['ID'] == id].tolist())[0]
            new_players = players.drop(player_index)
            new_players.to_excel("Player_db.xlsx")
    # search and delete from users data base
    users = pd.read_excel('Users_db.xlsx')
    for user_id in users['ID']:
        if id == user_id:
            user_index = (users.index[users['ID'] == id].tolist())[0]
            new_users = users.drop(user_index)
            new_users.to_excel("Users_db.xlsx")
            print('User deleted!')
            return 0
    # if the id wasn't found in any data base
    print('Error - ID not found')
    return 1


def Delete_Question():
    """ function print all questions in the relevant category , and deletes the question the user chose"""
    while True:
        category = input('Please choose a category to delete from (School, Home or Public Places): ')
        if category == 'Home' or category == 'School' or category == 'Public Places':
            break
        else:
            print('Invalid category!')
    questions = pd.read_excel('Question_db_new.xlsx')
    # making an index list of questions in the category
    q_list = questions.index[questions['Category'] == category].tolist()
    print('Please choose a question to delete:')
    index = 1
    # printing questions in the category
    for q_index in q_list:
        print('{0}: {1}'.format(index, questions.loc[q_index]['Question']))
        index += 1
    q_to_delete = int(input())
    print('You chose to delete the question: {0}'.format(questions.loc[q_list[q_to_delete - 1]]['Question']))
    # delete question and update data base
    new_questions = questions.drop(q_list[q_to_delete - 1])
    new_questions.to_excel("Question_db_new.xlsx")
    print('Question deleted!')


def Add_Question():
    """ function adds question according to the category the user chose"""
    while True:
        category = input('Please choose category to add a question (School, Home or Public Places): ')
        if category == 'Home' or category == 'School' or category == 'Public Places':
            break
        else:
            print('Invalid category!')
    ques = input('Please enter the question to add: ')
    print('Please enter 3 possible answers: ')
    right_answer = input('The right answer: ')
    answer2 = input('Answer 2: ')
    answer3 = input('Answer 3: ')
    questions = pd.read_excel('Question_db_new.xlsx')
    new_questions = questions.append(
        {'Category': category, 'Question': ques, 'Answer_A': right_answer, 'Answer_B': answer2, 'Answer_C': answer3,
         'Mistakes': 0, 'Right Answer': right_answer}, ignore_index=True)
    new_questions.to_excel("Question_db_new.xlsx")
    print('Question added!')
    if ques in new_questions.Question.values:
        return 0
    return 1


def Reset_Player():
    """ function deletes player's grades and last game data"""
    id = int(input('Please enter ID of the child to reset data: '))
    players = pd.read_excel('Player_db.xlsx')
    if not (id in players.ID.values):
        print('ID not found')
        return 1
    # reset data in players data base
    index = players.index[players['ID'] == id].tolist()[0]
    players.at[index, 'Login_count'] = 0
    players.at[index, ('Last_Login', 'Q1', 'A1', 'Q2', 'A2', 'Q3', 'A3', 'Q4', 'A4', 'Q5', 'A5', 'Last grade')] = "NaN"
    players.to_excel('Player_db.xlsx')
    id_db = '{0}.xlsx'.format(id)
    # create an empty data frame and overwrite player's data base
    empty_db = pd.DataFrame(columns=['Date', 'Grade'])
    empty_db.to_excel(id_db)
    print('Players data was Reset')
    return 0


# -----------------------------------


def login_report(id):
    """Print the player last login times"""
    playerDB = pd.read_excel('Player_db.xlsx')
    index = 0
    for kid in playerDB['ID']:
        if int(kid) == int(id):
            print("The player last login was at:", playerDB['Last_Login'][index])
            return 0
        index += 1
    return 1


def Print_Last_Mistake(id):
    """Print the last game mistakes"""
    playerDB = pd.read_excel('Player_db.xlsx')
    question = pd.read_excel('Question_db_new.xlsx')
    index = 0
    QandA = []
    for kid in playerDB['ID']:
        # QandA=[('Q1','A1'),('Q2','A2'),('Q3','A3'),('Q4','A4'),('Q5','A5')]
        if int(kid) == int(id):
            for n in range(1, 6):
                a = 'A' + str(n)
                q = 'Q' + str(n)
                QandA.append((playerDB[q][index], playerDB[a][index]))
        index += 1
    for index in range(0, len(QandA)):
        i = 0
        for q in question['Question']:
            if str(q) == str(QandA[index][0]):
                if not (question['Right Answer'][i] == QandA[index][1]):
                    if QandA[index][1] == 's':
                        print("Question:\n{0}\nwas skipped.\nThe correct answer is:\n{1}".
                              format(QandA[index][0], question['Right Answer'][i]))
                    else:
                        print("The question:\n{0}\nis incorrect.\nYour answer:\n{1}\nThe correct answer is:\n{2}"
                              .format(QandA[index][0], QandA[index][1], question['Right Answer'][i]))
                    print()
                break
            i += 1


def Print_Last_Grade(id):
    playerDB = pd.read_excel('Player_db.xlsx')
    index = 0
    for kid in playerDB['ID']:
        if int(kid) == int(id):
            print(playerDB['Last grade'][index])
            return 0
        index += 1
    return 1


def View_All(user_type):
    """View all the user type"""
    # Need to change to the right file path
    UsersDB = pd.read_excel('Users_db.xlsx')
    index = 0
    for user in UsersDB['ID']:
        types = int(UsersDB['Type'][index])
        if types == user_type:
            print(user)
        index = index + 1


def instructions():
    file = open("instruction1.txt", 'r')
    print(file.read())


def Calculate_Grade(id):
    """ Calculate the kid (id) last game grade """
    playerDB = pd.read_excel('Player_db.xlsx')
    QuestionsDB = pd.read_excel('Question_db_new.xlsx')
    index, i, score = 0, 0, 0
    for kid in playerDB['ID']:
        if kid == id:
            index = i
            Q, A = 'Q', 'A'
            # Found the kid's id
            for j in range(1, 6):
                Q, A = 'Q', 'A'
                Q = Q + str(j)
                A = A + str(j)
                Qindex = 0
                for QST in QuestionsDB['Question']:
                    # playerDB[Q][index]
                    if QST == playerDB[Q][index]:
                        # This is the question index
                        if playerDB[A][index] == QuestionsDB['Right Answer'][Qindex]:
                            # Correct answer
                            score += 20
                    Qindex += 1

            gradeFile = pd.read_excel('{0}.xlsx'.format(id))
            NewGrade = gradeFile.append(
                {'Grade': score, 'Date': time.asctime(time.localtime(time.time()))}, ignore_index=True)
            NewGrade.to_excel("{0}.xlsx".format(id))
        i += 1
    return score


def Game(category, id):
    # Loading the data bases
    Q_and_A_write = load_workbook(filename="Player_db.xlsx")
    sheet = Q_and_A_write.active
    Answer_write = pd.read_excel("Player_db.xlsx", "Sheet1")
    Q_and_A_read = pd.read_excel("Question_db_new.xlsx", "Sheet1")

    # Generate random questions index for each category
    # School category
    if category == 1:
        num = 0
        Id_Index = 0
        Q_Index_Arr = []
        for cat in Q_and_A_read['Category']:
            if cat == 'School':
                num += 1
                Q_Index_Arr.append(Q_and_A_read['Index'][Id_Index])
            Id_Index += 1
        Q_Random_Index_Arr = sample(Q_Index_Arr, 5)

    # Home category
    elif category == 2:
        num = 0
        Id_Index = 0
        Q_Index_Arr = []
        for cat in Q_and_A_read['Category']:
            if cat == 'Home':
                num += 1
                Q_Index_Arr.append(Q_and_A_read['Index'][Id_Index])
            Id_Index += 1
        Q_Random_Index_Arr = sample(Q_Index_Arr, 5)

    # Public places category
    elif category == 3:
        num = 0
        Id_Index = 0
        Q_Index_Arr = []
        for cat in Q_and_A_read['Category']:
            if cat == 'Public Places':
                num += 1
                Q_Index_Arr.append(Q_and_A_read['Index'][Id_Index])
            Id_Index += 1
        Q_Random_Index_Arr = sample(Q_Index_Arr, 5)

    # From all the category's
    elif category == 4:
        num = 0
        for cat in Q_and_A_read['Category']:
            num += 1
        Q_Random_Index_Arr = sample(range(1, num + 1), 5)

    # Give back the line number for the player id
    Id_Index = 0
    Ind = 0
    for Id in Answer_write['ID']:
        if int(id) == int(Id):
            Id_Index = Ind
        Ind += 1

    Q_Arr = []
    A_Arr = []

    # Insert the questions to the Q_Arr
    Q_Arr.append(Q_and_A_read['Question'][Q_Random_Index_Arr[0] - 1])
    Q_Arr.append(Q_and_A_read['Question'][Q_Random_Index_Arr[1] - 1])
    Q_Arr.append(Q_and_A_read['Question'][Q_Random_Index_Arr[2] - 1])
    Q_Arr.append(Q_and_A_read['Question'][Q_Random_Index_Arr[3] - 1])
    Q_Arr.append(Q_and_A_read['Question'][Q_Random_Index_Arr[4] - 1])

    # Enter and print the answers to the A_Arr
    for answer in range(0, 5):
        A_Arr.append(Q_and_A_read['Answer_A'][Q_Random_Index_Arr[answer] - 1])
        A_Arr.append(Q_and_A_read['Answer_B'][Q_Random_Index_Arr[answer] - 1])
        A_Arr.append(Q_and_A_read['Answer_C'][Q_Random_Index_Arr[answer] - 1])

    # Give back the column number for column Q1
    In_col = 0
    for col in Answer_write:
        if 'Q1' == str(col):
            Id_In = In_col
        In_col += 1

    # Save the questions in the player db and the answers as NaN
    Q1_Cell = sheet.cell(row=Id_Index + 2, column=Id_In + 1)
    Q1_Cell.value = Q_Arr[0]
    Q2_Cell = sheet.cell(row=Id_Index + 2, column=Id_In + 3)
    Q2_Cell.value = Q_Arr[1]
    Q3_Cell = sheet.cell(row=Id_Index + 2, column=Id_In + 5)
    Q3_Cell.value = Q_Arr[2]
    Q4_Cell = sheet.cell(row=Id_Index + 2, column=Id_In + 7)
    Q4_Cell.value = Q_Arr[3]
    Q5_Cell = sheet.cell(row=Id_Index + 2, column=Id_In + 9)
    Q5_Cell.value = Q_Arr[4]
    A1_Cell = sheet.cell(row=Id_Index + 2, column=Id_In + 2)
    A1_Cell.value = 'NaN'
    A2_Cell = sheet.cell(row=Id_Index + 2, column=Id_In + 4)
    A2_Cell.value = 'NaN'
    A3_Cell = sheet.cell(row=Id_Index + 2, column=Id_In + 6)
    A3_Cell.value = 'NaN'
    A4_Cell = sheet.cell(row=Id_Index + 2, column=Id_In + 8)
    A4_Cell.value = 'NaN'
    A5_Cell = sheet.cell(row=Id_Index + 2, column=Id_In + 10)
    A5_Cell.value = 'NaN'
    Q_and_A_write.save(filename="Player_db.xlsx")

    # a while function for the game
    flag = True
    Q_Num = 1
    while flag:

        # Correcting the question number and column index
        if Q_Num == 6:
            Q_Num = 1
        if Q_Num == 0:
            Q_Num = 5
        if Q_Num == 1:
            arg = 2
        if Q_Num == 2:
            arg = 4
        if Q_Num == 3:
            arg = 6
        if Q_Num == 4:
            arg = 8
        if Q_Num == 5:
            arg = 10

        # Print the questions and the answers
        print("\nQuestion number ", Q_Num, ": ", Q_Arr[Q_Num - 1])
        print("\nAnswer number 1: ", A_Arr[(Q_Num - 1) * 3])
        print("Answer number 2: ", A_Arr[((Q_Num - 1) * 3) + 1])
        print("Answer number 3: ", A_Arr[((Q_Num - 1) * 3) + 2])
        check_if_answer = sheet.cell(row=Id_Index + 2, column=Id_In + arg)

        # Print an answer to an answered question
        if check_if_answer.value != 'NaN':
            print("\nYou hve answer to this question: ", check_if_answer.value)

        # Print the options for each question
        ans = str(input("\nPress 1 to choose answer number 1 \nPress 2 to choose answer number 2 \nPress 3 to choose "
                        "answer number 3 \nPress 4 to skip the question \nPress 5 to move to the next question game "
                        "\nPress 6 to move to the previous question \nPress 7 to end a finished or an unfinished \n"))
        while ans != '1' and ans != '2' and ans != '3' and ans != '4' and ans != '5' and ans != '6' and ans != '7':
            ans = str(input("Wrong input, try again: "))

        # let the player answer a question
        if ans == '1':
            Answer_Cell = sheet.cell(row=Id_Index + 2, column=Id_In + arg)
            Answer_Cell.value = A_Arr[(Q_Num - 1) * 3]
            Q_and_A_write.save(filename="Player_db.xlsx")
            Q_Num += 1
        if ans == '2':
            Answer_Cell = sheet.cell(row=Id_Index + 2, column=Id_In + arg)
            Answer_Cell.value = A_Arr[((Q_Num - 1) * 3) + 1]
            Q_and_A_write.save(filename="Player_db.xlsx")
            Q_Num += 1
        if ans == '3':
            Answer_Cell = sheet.cell(row=Id_Index + 2, column=Id_In + arg)
            Answer_Cell.value = A_Arr[((Q_Num - 1) * 3) + 2]
            Q_and_A_write.save(filename="Player_db.xlsx")
            Q_Num += 1
        if ans == '4':
            Answer_Cell = sheet.cell(row=Id_Index + 2, column=Id_In + arg)
            Answer_Cell.value = 's'
            Q_and_A_write.save(filename="Player_db.xlsx")
            Q_Num += 1
        if ans == '5':
            Q_Num += 1
        if ans == '6':
            Q_Num -= 1
        if ans == '7':
            Score = Calculate_Grade(id)
            see = str(input("\nto see the grade press 1 \nto exit the game back to the menu press 2\n"))
            while see != '1' and see != '2':
                see = str(input("Wrong input, try again: "))
            if see == '1':
                Grade_Cell = sheet.cell(row=Id_Index + 2, column=Id_In + 11)
                Grade_Cell.value = Score
                Q_and_A_write.save(filename="Player_db.xlsx")
                print("Your score is: ", Grade_Cell.value)
                break
            elif see == '2':
                break


Game(4, 311369318)


def Choose_Category(id):
    print('Choose game category')
    print('1- School')
    print('2- Home')
    print('3- Public places')
    print('4- Random questions')
    choice = int(input())
    while choice != 1 and choice != 2 and choice != 3 and choice != 4:
        print('Wrong input, try again')
        print('1- School')
        print('2- Home')
        print('3- Public places')
        print('4- Random questions')
        choice = int(input())
    Game(choice, id)


def Print_Last_Game(id):
    # Loading the data bases
    Player_db = pd.read_excel("Player_db.xlsx", "Sheet1")

    # Check if the id exist in the DB and if exist save the index line
    flag = True
    index = 0
    i = 0
    for Id in Player_db['ID']:
        if int(id) == int(Id):
            flag = False
            index = i
        i += 1
    if not flag:

        # Check if the player did a test or not
        if str(Player_db['Q1'][index]) == str('nan'):
            print("The player didn't play yet")
        else:

            # Print the player last game
            print("question 1: ", Player_db['Q1'][index])
            if Player_db['A1'][index] == 's':
                print("the player skipped the question")
            else:
                print("answer 1: ", Player_db['A1'][index])
            print("question 2: ", Player_db['Q2'][index])
            if Player_db['A2'][index] == 's':
                print("the player skipped the question")
            else:
                print("answer 2: ", Player_db['A2'][index])
            print("question 3: ", Player_db['Q3'][index])
            if Player_db['A3'][index] == 's':
                print("the player skipped the question")
            else:
                print("answer 3: ", Player_db['A3'][index])
            print("question 4: ", Player_db['Q4'][index])
            if Player_db['A4'][index] == 's':
                print("the player skipped the question")
            else:
                print("answer 4: ", Player_db['A4'][index])
            print("question 5: ", Player_db['Q5'][index])
            if Player_db['A5'][index] == 's':
                print("the player skipped the question")
            else:
                print("answer 5: ", Player_db['A5'][index])
    else:
        print("Player ID was not found")


def Player_Menu(id):
    print()
    print("Choose an option: ")
    print("1- Play game \n2- Show game instructions \n3- Show grades")
    print("4- Show last played game \n5- Show last game skipped question \n6- Show the latest grade")
    print("7- Exit to login screen")

    choice = int(input())
    if choice == 1:
        Choose_Category(id)
        Player_Menu(id)
    if choice == 2:
        instructions()
        Player_Menu(id)
    if choice == 3:
        Print_Grades(id)
        Player_Menu(id)
    if choice == 4:
        Print_Last_Game(id)
        Player_Menu(id)
    if choice == 5:
        View_Skip(id)
        Player_Menu(id)
    if choice == 6:
        Print_Last_Grade(id)
        Player_Menu(id)
    if choice == 7:
        Login_And_SignUp()


# --------------------------------------------------


def Parent_Menu(id):
    print()
    print("Choose an option: ")
    print("1- Add kid \n2- View kid\n3- Show last grade")
    print("4- Show the kid's login count \n5- Show last game skipped question \n6- Play example game")
    print("7- Show kid's last game \n8- Show the kid's last game mistake \n9- Show the kid's last login date")
    print("10- Exit to login screen")

    choice = int(input())
    if choice == 1:
        Add_Kid(id)
        Parent_Menu(id)
    if choice == 2:
        View_Kid(id)
        Parent_Menu(id)
    if choice == 3:
        Id = int(input('Please enter child ID: '))
        Print_Grades(Id)
        Parent_Menu(id)
    if choice == 4:
        Print_Login_Count()
        Parent_Menu(id)
    if choice == 5:
        ID = int(input('Please enter child ID: '))
        View_Skip(ID)
        Parent_Menu(id)
    if choice == 6:
        Example_Game()
        Parent_Menu(id)
    if choice == 7:
        Id = int(input('Please enter child ID: '))
        Print_Last_Game(Id)
        Parent_Menu(id)
    if choice == 8:
        Id = int(input('Please enter child ID: '))
        Print_Last_Mistake(Id)
        Parent_Menu(id)
    if choice == 9:
        Id = int(input('Please enter child ID: '))
        login_report(Id)
        Parent_Menu(id)
    if choice == 10:
        Login_And_SignUp()


# --------------------------------------------------


def Professional_Menu(id):
    print()
    print('Choose an option: ')
    print("1- Reports\n2- Watch child's grades\n3- Watch child's last games skipped questions\n"
          "4- Reset players data\n5- Add a question\n6- Delete a question")
    print('7- Watch the most mistaken question\n8- Delete a user\n9- Exit to login screen')
    choice = int(input())
    if choice == 1:
        print('1- Players report\n2- Parent report')
        report = int(input())
        if report == 1:
            View_All(1)
            Professional_Menu(id)
        elif report == 2:
            View_All(2)
            Professional_Menu(id)
    if choice == 2:
        Id = int(input("Please enter child's ID: "))
        Print_Grades(Id)
        Professional_Menu(id)
    if choice == 3:
        Id = int(input("Please enter child's ID: "))
        View_Skip(Id)
        Professional_Menu(id)
    if choice == 4:
        Reset_Player()
        Professional_Menu(id)
    if choice == 5:
        Add_Question()
        Professional_Menu(id)
    if choice == 6:
        Delete_Question()
        Professional_Menu(id)
    if choice == 7:
        Most_Mistakes()
        Professional_Menu(id)
    if choice == 8:
        Delete_User()
        Professional_Menu(id)
    if choice == 9:
        Login_And_SignUp()


# --------------------------------------------------


# login and sign-up function
def Login_And_SignUp():
    # Loading the data bases
    write = load_workbook(filename="Users_db.xlsx")
    sheet = write.active
    Users_db = pd.read_excel("Users_db.xlsx", "Sheet1")
    login_count = load_workbook(filename="Player_db.xlsx")
    Login_c = login_count.active

    # Choose if you want to sign in or sign up or exit the system
    print("Welcome...")
    welcome = input("Press y to login \nPress n to sign-up \nPress any other key to exit the system \n")

    # Sign-up function
    if welcome == "n" or welcome == "N":

        # Check the index of the first free line in Users_DB
        count = 0
        for i in Users_db['ID']:
            count += 1
        count += 1
        while True:
            username = int(input("Enter a username: "))
            flag = False

            # Check if the id already exist in the DB
            for Id in Users_db['ID']:
                if username == Id:
                    flag = True
            if not flag:

                # Enter and save the new id, password and user type to DB
                IdCell = sheet.cell(row=count + 1, column=1)
                IdCell.value = username
                password = int(input("Enter a password: "))
                PasswordCell = sheet.cell(row=count + 1, column=2)
                PasswordCell.value = password
                print("Type 1 for Player")
                print("Type 2 for Parent")
                print("Type 3 for Professional")
                usertype = int(input("Enter the user type: "))
                while usertype != 1 and usertype != 2 and usertype != 3:
                    usertype = int(input("Wrong input, try again: "))
                TypeCell = sheet.cell(row=count + 1, column=3)
                TypeCell.value = usertype
                break
            print("ID already exist")
        write.save(filename="Users_db.xlsx")

        # Login and activate player menu
        if usertype == 1:

            # Open an excel file for each player
            d = {'Date': [], 'Grade': []}
            df = pd.DataFrame(data=d)
            user_db = '{0}.xlsx'.format(username)
            df.to_excel(user_db)
            i = 0

            # Check the index of the first free line in Player_DB
            for row in Login_c.rows:
                i = i + 1
                Line = i
            Line = Line + 1
            tempLine = Line

            # Save id, date index and login count to the player_DB and activate player menu
            CurDate = time.asctime(time.localtime(time.time()))
            NumCell = Login_c.cell(row=Line, column=1)
            NumCell.value = tempLine - 2
            IDcell = Login_c.cell(row=Line, column=2)
            IDcell.value = username
            DateCell = Login_c.cell(row=Line, column=3)
            DateCell.value = CurDate
            LogCell = Login_c.cell(row=Line, column=4)
            LogCell.value = 1
            login_count.save(filename="Player_db.xlsx")
            print("\nWelcome to the Player Menu")
            Player_Menu(username)

        # Login and activate parent menu
        elif usertype == 2:
            print("\nWelcome to the Parent Menu\n")
            Parent_Menu(username)

        # Login and activate professional menu
        elif usertype == 3:
            print("\nWelcome to the Professional Menu\n")
            Professional_Menu(username)

    # Login function
    elif welcome == "y" or welcome == "Y":
        while True:
            username = int(input("Enter a username: "))
            i = 0

            # Check the index line of the username in users_DB
            for row in sheet.rows:
                i = i + 1
                for cell in row:
                    if cell.value == username:
                        line = i

            # Check if the id exist in the DB
            flag = True
            for Id in Users_db['ID']:
                if username == Id:
                    flag = False
            if not flag:
                while True:

                    # Compare the password in the DB to the password that the user enter
                    password = int(input("Enter a password: "))
                    flag = False
                    if password != Users_db['Password'][line - 2]:
                        flag = True
                    if not flag:

                        # Login and activate player menu
                        if Users_db['Type'][line - 2] == 1:

                            # Check the index line of the username in Player_DB
                            k = 0
                            LINE = 0
                            for row in Login_c.rows:
                                for cell in row:
                                    if cell.value == username:
                                        LINE = k
                                k = k + 1

                            # Update date and login count to the player_DB and activate player menu
                            CurDate = time.asctime(time.localtime(time.time()))
                            DateCell = Login_c.cell(row=LINE + 1, column=3)
                            DateCell.value = CurDate
                            LogCell = Login_c.cell(row=LINE + 1, column=4)
                            LogCell.value = LogCell.value + 1
                            login_count.save(filename="Player_db.xlsx")
                            print("\nWelcome to the Player Menu\n")
                            return Player_Menu(username)

                        # Login and activate parent menu
                        elif Users_db['Type'][line - 2] == 2:
                            print("\nWelcome to the Parent Menu\n")
                            return Parent_Menu(username)

                        # Login and activate professional menu
                        elif Users_db['Type'][line - 2] == 3:
                            print("\nWelcome to the Professional Menu\n")
                            return Professional_Menu(username)
                    print("Wrong password, try again")
                break
            print("ID not exist in the system")


# Login_And_SignUp()
